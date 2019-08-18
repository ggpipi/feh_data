#!/usr/bin/env python
# -*- coding: utf-8 -*-
import lxml
import requests  # 导入网页请求库
from bs4 import BeautifulSoup  # 导入网页解析库
import time
import random
import pickle
import sys
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import http.client
import hashlib #md5替换
import urllib
import random
import requests
import json

def creaturl(txt, yuan='jp',mubiao='zh'):  # 接收三个字符串参数
    appid = 'xxxxxxxxxxxxx' #你的appid
    secretKey = 'xxxxxxxxxxxxxxx' #你的密钥
    httpClient = None
    myurl = '/api/trans/vip/translate'
    q = txt
    fromLang = yuan
    toLang = mubiao
    salt = random.randint(32768, 65536)

    sign = appid+q+str(salt)+secretKey
    #对签名进行加密
    m2=hashlib.md5(sign.encode(encoding='UTF-8')).hexdigest()
    myurl = myurl+'?appid='+appid+'&q='+urllib.parse.quote(q)+'&from='+fromLang+'&to='+toLang+'&salt='+str(salt)+'&sign='+m2
    return myurl

def requ(myurl):
    try:

        re = requests.get("https://fanyi-api.baidu.com"+myurl)
        dic = json.loads(re.text)
        # print("翻译结果如下：")
        ##re.text是字符串类型，利用json模块中的loads函数把它转换为字典
        print('ok')
        return dic["trans_result"][0]["dst"]
    except Exception as e:
        print(e)

sys.setrecursionlimit(100000)  # 你想设置的递归深度（可为任意值）

class Hero:
    def __init__(self, url, name, tcn, rank, color, weapon, move, rare, xingge=None, character=None, article1=None, article11 = None, article2 = None, article22 = None):
        self.url = url
        self.name = name
        self.tcn = tcn
        self.rank = rank
        self.color = color
        self.weapon = weapon
        self.move = move
        self.rare = rare
        self.xingge = xingge
        self.character = character
        self.article1 = article1
        self.article11 = article11
        self.article2 = article2
        self.article22 = article22

def asad(pa = False, make_dict = False):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36"}

    # 开启一个session会话
    session = requests.session()

    # 设置请求头信息
    session.headers = headers

    if make_dict != False:
        hero_dict = {}
        url_dict = 'https://feheroes.gamepedia.com/'
        list_hero = 'List_of_Heroes'
        res = session.get(url=url_dict+list_hero)
        text = res.content
        soup = BeautifulSoup(text, 'lxml')
        list_h = soup.find('tbody')
        list_h = list_h.find_all('tr', class_='hero-filter-element')
        for i in list_h:
            url1 = i.find('a')['href']
            url1 = url_dict + 'Heroic_Ordeals:_' + url1[1:] + '%27s_Trial'
            res = session.get(url = url1)
            text = res.content
            soup = BeautifulSoup(text, 'lxml')
            jp = soup.find(text='Japanese\n')
            tc = soup.find(text='Traditional Chinese (Taiwan)\n')
            try:
                jpn = jp.next_element.next_element.text
                tcn = tc.next_element.next_element.text
                jpn = jpn.split('\u3000')[-1].strip()
                tcn = tcn.split('\u3000')[-1].strip()
                print(jpn+'\t'+tcn)
                hero_dict[jpn] = tcn
            except:
                print('pass')
            time.sleep(0.5)
        print(len(hero_dict.keys()))
        with open('hero_dict_game8.pkl', 'wb')as f:
            pickle.dump(hero_dict, f)
        print("make dict finish!")
    else:
        with open('hero_dict_game8.pkl', 'rb')as f:
            hero_dict = pickle.load(f)
        print(len(hero_dict.keys()))
        print('load dict finish')

    if pa != False:
        wb = Workbook()  # 创建文件对象

        # grab the active worksheet
        ws = wb.active  # 获取第一个sheet

        url = r'https://game8.jp/fe-heroes/116312'

        res1 = session.get(url=url)
        text = res1.content
        soup = BeautifulSoup(text, 'lxml')
        list1 = soup.find('table', class_='a-table a-table a-table tablesorter')
        list2 = list1.find('tbody')
        list3 = list2.find_all('tr')
        listA = []

        with open('game8.txt', 'w', encoding='utf-8')as g:
            a = 1
            for i in list3:
                time1 = time.time()
                url2 = i.find('a')['href']
                print(url2)
                img_url = i.find('img')['src']
                img = session.get(img_url)


                res2 = session.get(url=url2)
                text2 = res2.content
                soup2 = BeautifulSoup(text2, 'lxml')
                show = soup2.find('div', class_='archive-style-wrapper')
                name = show.find('h2').contents[0]
                name = name.replace('の評価と基本情報', '')
                tcn = ''
                print(name)
                if name in hero_dict.keys():
                    tcn = hero_dict[name]
                    print('1'+tcn)
                else:
                    for j in hero_dict:
                        if j in name:
                            tcn = hero_dict[j]
                            print('2'+tcn)
                            break

                with open('./pic/'+name+'.jpeg', 'wb')as f1:
                    f1.write(img.content)

                rank = show.find('span', style='font-size:140%;').text
                print(rank)

                cwm = show.find(text='属性/武器/移動')
                cwm = cwm.next_element.next_element

                # list4 = show.find_all('table', class_='a-table a-table a-table ', limit=5)

                # rank = list4[0].find('span').contents[0]
                # canshu = show.find('div', class_='align')
                canshu = cwm.find_all('a', class_='a-link')
                color = str(canshu[0].contents[1])
                weapon = str(canshu[1].contents[1])[-1]
                move = str(canshu[2].contents[1])
                rare = show.find(text='排出レアリティ')
                rare = str(rare.next_element.next_element.find('a').contents[1])

                xingge = [' ', ' ', ' ', ' ', ' ']
                character = ' '
                article1 = ' '
                article2 = ' '
                article11 = ' '
                article22 = ' '
                try:
                    osusume = show.find(text='おすすめの個体値').next_element.next_element

                    list5 = osusume.find_all('tr')
                    for j in range(1, 6):
                        xingge[j - 1] = list5[j].find_all('td')[2].contents[0].strip()

                    character = show.find('h4', class_='a-header--4', id='hs_2').contents[0]
                    character = character.replace(name + 'は', '')
                    character = character.replace('がおすすめ', '')
                    article = show.find_all('p', class_='a-paragraph', limit=6)
                    article1 = article[4].text
                    my = creaturl(article1)
                    article11 = requ(my)

                    article2 = article[5].text
                    if article2 == '':
                        article2 = '_None'
                    else:
                        time.sleep(1.5)
                        my = creaturl(article2)
                        article22 = requ(my)
                except:
                    pass

                ws['A' + str(a)] = name
                ws['B' + str(a)] = tcn
                img = Image('./pic/'+name+'.jpeg')
                img.width, img.height = (50, 50)
                ws.add_image(img, 'C'+str(a))
                ws['D' + str(a)] = rank
                ws['E' + str(a)] = color
                ws['F' + str(a)] = weapon
                ws['G' + str(a)] = move
                ws['H' + str(a)] = rare
                ws['I' + str(a)] = xingge[0]
                ws['J' + str(a)] = xingge[1]
                ws['K' + str(a)] = xingge[2]
                ws['L' + str(a)] = xingge[3]
                ws['M' + str(a)] = xingge[4]
                ws['N' + str(a)] = character
                ws['O' + str(a)] = article2
                ws['P' + str(a)] = article22
                ws['Q' + str(a)] = article1
                ws['R' + str(a)] = article11
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 6
                ws.column_dimensions['D'].width = 7
                ws.column_dimensions['E'].width = 9
                ws.column_dimensions['F'].width = 6
                ws.column_dimensions['G'].width = 7
                ws.column_dimensions['H'].width = 7
                ws.column_dimensions['O'].width = 30
                ws.column_dimensions['P'].width = 30
                ws.column_dimensions['Q'].width = 30
                ws.column_dimensions['R'].width = 30
                ws.column_dimensions['N'].width = 17
                ws.row_dimensions[a].height = 40
                a = a+1

                listA.append(Hero(url2, name, tcn, rank, color, weapon, move, rare, xingge, character, article1, article11, article2, article22))
                time.sleep(1.1)
                g.write(
                    name + '\t' + tcn + '\t' + rank + '\t' + color + '\t' + weapon + '\t' + move + '\t' + rare + '\t' + url2 + '\t' + '\t'.join(
                        xingge) + '\t' + character + '\t' + article2 + '\t' + article22 + '\t' + article1 + '\t' + article11 + '\n')
                print(time.time() - time1)
                # if a == 4:
                #     break

        with open('game8.pkl', 'wb') as f:
            pickle.dump(listA, f)
        wb.save(filename='game8.xlsx')
        return listA
    else:
        with open('game8.pkl', 'rb') as f:
            doc = pickle.load(f)
        return doc

listb = asad(pa=True, make_dict=False)
a = 1

